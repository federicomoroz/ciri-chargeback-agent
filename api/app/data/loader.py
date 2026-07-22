"""
Excel → SQLite loader.

Dataset quirks:
- Row 1: decorative merged title (skip)
- Row 2: real headers (used as column mapping)
- Data starts at Row 3
- Sheet names contain emojis → use substring matching
- All dates/timestamps are strings (openpyxl reads them as str)
- Logs.Codigo is a string, not int
- Policies "Politica" column: "POL-XXX-NNN — Description" (em-dash)
"""

import sqlite3
from datetime import datetime, timezone

import openpyxl


def _find_sheet(wb: openpyxl.Workbook, keyword: str) -> str:
    """Find sheet by case-insensitive keyword substring (handles emoji prefixes)."""
    for name in wb.sheetnames:
        if keyword.lower() in name.lower():
            return name
    raise ValueError(f"Sheet containing '{keyword}' not found. Available: {wb.sheetnames}")


def _parse_policy_field(raw: str) -> tuple[str, str]:
    """'POL-FRD-001 — Score minimo' → ('POL-FRD-001', 'Score minimo')"""
    if not raw:
        return ("", "")
    # Try em-dash with spaces (U+2014)
    if " \u2014 " in raw:
        parts = raw.split(" \u2014 ", maxsplit=1)
        return parts[0].strip(), parts[1].strip()
    # Try regular dash with spaces
    if " - " in raw:
        parts = raw.split(" - ", maxsplit=1)
        return parts[0].strip(), parts[1].strip()
    # Fallback: treat whole thing as code
    return raw.strip(), ""


def _parse_sheet(wb: openpyxl.Workbook, keyword: str, columns: list[str]) -> list[dict]:
    """Generic parser: find sheet by keyword, skip row 1, row 2=headers (ignored), data from row 3."""
    sheet_name = _find_sheet(wb, keyword)
    ws = wb[sheet_name]
    rows = []
    for row_num in range(3, ws.max_row + 1):
        values = [ws.cell(row=row_num, column=c).value for c in range(1, len(columns) + 1)]
        if all(v is None for v in values):
            continue
        row_dict = dict(zip(columns, values))

        # Type coercions
        if "amount_usd" in row_dict and row_dict["amount_usd"] is not None:
            row_dict["amount_usd"] = float(row_dict["amount_usd"])
        if "fraud_score" in row_dict and row_dict["fraud_score"] is not None:
            row_dict["fraud_score"] = int(row_dict["fraud_score"])
        if "resolution_days" in row_dict and row_dict["resolution_days"] is not None:
            row_dict["resolution_days"] = int(row_dict["resolution_days"])
        # Ensure date/timestamp fields are strings
        for date_field in ("date", "open_date", "close_date", "timestamp"):
            if date_field in row_dict and row_dict[date_field] is not None:
                row_dict[date_field] = str(row_dict[date_field])
        # HTTP code must be string
        if "code" in row_dict and row_dict["code"] is not None:
            row_dict["code"] = str(int(float(str(row_dict["code"])))) if row_dict["code"] != "" else "0"

        rows.append(row_dict)
    return rows


def load_excel(file_path: str) -> dict:
    """Load all 4 data sheets from the Excel file (skips README sheet).

    Returns:
        {
            'transactions': [...],  # 100 rows
            'cases': [...],         # 60 rows
            'policies': [...],      # 17 rows
            'logs': [...],          # 150 rows
        }
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    transactions = _parse_sheet(wb, "Transacciones", [
        "id", "client_id", "merchant", "amount_usd", "date",
        "payment_method", "country", "channel", "device",
        "fraud_score", "status", "notes",
    ])

    raw_cases = _parse_sheet(wb, "Hist", [
        "case_id", "transaction_id", "motivo", "resolution",
        "resolution_days", "analyst", "observations",
        "open_date", "close_date",
    ])

    raw_policies = _parse_sheet(wb, "Pol", [
        "category", "politica_raw", "description", "reference",
    ])

    logs = _parse_sheet(wb, "Logs", [
        "timestamp", "transaction_id", "event", "service",
        "code", "detail", "severity",
    ])

    # Parse policy code + name from compound "Politica" field
    policies = []
    for p in raw_policies:
        code, name = _parse_policy_field(p.get("politica_raw") or "")
        policies.append({
            "code": code,
            "name": name,
            "category": p.get("category", ""),
            "description": p.get("description", ""),
            "reference": p.get("reference", ""),
        })

    return {
        "transactions": transactions,
        "cases": raw_cases,
        "policies": policies,
        "logs": logs,
    }


def init_sqlite(db_path: str, data: dict) -> None:
    """Create all tables and insert data. Idempotent (INSERT OR IGNORE)."""
    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")

    conn.execute("""
        CREATE TABLE IF NOT EXISTS transactions (
            id TEXT PRIMARY KEY,
            client_id TEXT NOT NULL,
            merchant TEXT NOT NULL,
            amount_usd REAL NOT NULL,
            date TEXT NOT NULL,
            payment_method TEXT NOT NULL,
            country TEXT NOT NULL,
            channel TEXT NOT NULL,
            device TEXT NOT NULL,
            fraud_score INTEGER NOT NULL,
            status TEXT NOT NULL,
            notes TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS cases (
            case_id TEXT PRIMARY KEY,
            transaction_id TEXT NOT NULL,
            motivo TEXT NOT NULL,
            resolution TEXT NOT NULL,
            resolution_days INTEGER NOT NULL,
            analyst TEXT NOT NULL,
            observations TEXT,
            open_date TEXT NOT NULL,
            close_date TEXT NOT NULL
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS policies (
            code TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            description TEXT NOT NULL,
            reference TEXT NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            transaction_id TEXT NOT NULL,
            event TEXT NOT NULL,
            service TEXT NOT NULL,
            code TEXT NOT NULL,
            detail TEXT NOT NULL,
            severity TEXT NOT NULL
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            transaction_id TEXT NOT NULL,
            analyst_decision TEXT NOT NULL,
            analyst_notes TEXT NOT NULL,
            final_outcome TEXT NOT NULL,
            judge_score REAL NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    # Indexes for common lookups
    conn.execute("CREATE INDEX IF NOT EXISTS idx_logs_tx ON logs(transaction_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_cases_tx ON cases(transaction_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tx_client ON transactions(client_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_tx_merchant ON transactions(merchant)")

    now = datetime.now(timezone.utc).isoformat()

    # Transactions
    conn.executemany(
        "INSERT OR IGNORE INTO transactions VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
        [(
            t["id"], t["client_id"], t["merchant"], t["amount_usd"],
            t["date"], t["payment_method"], t["country"], t["channel"],
            t["device"], t["fraud_score"], t["status"], t.get("notes"),
        ) for t in data["transactions"]],
    )

    # Cases
    conn.executemany(
        "INSERT OR IGNORE INTO cases VALUES (?,?,?,?,?,?,?,?,?)",
        [(
            c["case_id"], c["transaction_id"], c["motivo"], c["resolution"],
            c["resolution_days"], c["analyst"], c.get("observations"),
            c["open_date"], c["close_date"],
        ) for c in data["cases"]],
    )

    # Policies
    conn.executemany(
        "INSERT OR IGNORE INTO policies VALUES (?,?,?,?,?,?,?)",
        [(
            p["code"], p["name"], p["category"],
            p["description"], p["reference"], now, now,
        ) for p in data["policies"]],
    )

    # Logs
    conn.executemany(
        "INSERT INTO logs (timestamp, transaction_id, event, service, code, detail, severity) VALUES (?,?,?,?,?,?,?)",
        [(
            l["timestamp"], l["transaction_id"], l["event"],
            l["service"], l["code"], l["detail"], l["severity"],
        ) for l in data["logs"]],
    )

    conn.commit()
    conn.close()
